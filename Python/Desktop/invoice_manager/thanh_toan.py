from tkinter import *
import xlsxwriter
import openpyxl
from tkinter import messagebox
import obj
drink = ['Trà sữa', 'Trà đào', 'Trà thanh đào', 'Cà phê phin', 'Cà phê sữa', 'Cà phê đen', 'Bạc sỉu', 'Trà vải']
phep_tinh=""
branch_name = 'Branch Phan Van Tri'
#tạo option list
optionList = {0,1,2,3,4,5,6,7,8,9,10}
'''window = Tk()
q1 = IntVar()
q1.set(0)
q2 = IntVar()
q2.set(0)
q3 = IntVar()
q3.set(0)
q4 = IntVar()
q4.set(0)
q5 = IntVar()
q5.set(0)
q6 = IntVar()
q6.set(0)
q7 = IntVar()
q7.set(0)
q8 = IntVar()
q8.set(0)'''
def nhan_phim(ki_tu):
    global phep_tinh
    phep_tinh += str(ki_tu)
    display.configure(text=phep_tinh)
def xoa():
    global phep_tinh
    phep_tinh=""
    display.configure(text=phep_tinh)

def nhan_bang():
    # global phep_tinh
    try:
        dap_so = str(eval(phep_tinh))
        display.configure(text=dap_so)
        phep_tinh=""
    except:
        display.configure(text="Lỗi, vui lòng nhấn AC và nhập lại")
        phep_tinh=""
def add_invoice():
    msg = ''
    xlsx_content = []
    for i in range(8):
        if quan[i].get() != 0:
            msg += drink[i] + ' x ' + str(quan[i].get()) + '\n'
            line = [drink[i], str(quan[i].get())]
            xlsx_content.append(line)
    if msg != '':
        confirm = messagebox.askquestion(title='Confirm', message=msg, icon='info')
        if confirm == 'yes':
            add_to_excel(xlsx_content)
    else:
        messagebox.showwarning('Warning', 'Empty invoice!')
def add_to_excel(xlsx_content):
    # Create a workbook and add a worksheet.
    workbook = openpyxl.load_workbook('Expenses01.xlsx')
    # workbook = xlsxwriter.Workbook('Expenses01.xlsx')
    work_sheet = workbook['Sheet1']
    other = workbook['Sheet2']
    row = int(other.cell(1, 1).value)
    no = int(other.cell(1, 2).value)
    # Start from the first cell. Rows and columns are zero indexed.
    col = 1
    # Iterate over the data and write it out row by row.
    for line in (xlsx_content):
        print(line[0] + line[1])
        work_sheet.cell(row, col, value = no)
        work_sheet.cell(row, col + 1, value = line[0])
        work_sheet.cell(row, col + 2, value = line[1])
        work_sheet.cell(row, col + 3, value = branch_name)
        #work_sheet.write(row, col, line[0])
        #work_sheet.write(row, col + 1, line[1])
        row += 1
    other.cell(1, 1, value = row)
    other.cell(1, 2, value = no + 1)
    workbook.save('Expenses01.xlsx')
    workbook.close()
def create():
    global window
    window = Tk()
    global quan
    quan = []
    for i in range(8):
        temp = IntVar(window)
        temp.set(0)
        quan.append(temp)
    window.title('Main cho nhân viên')
    window.configure(background='blue')
    window.geometry('1200x600')
    #tạo thanh tính
    display = Label(window, width=45, height=3)
    display.grid(row=0, column=0, columnspan=3)
    #tao các nút
    bt1 = Button(window, text=drink[0], width=15, height=3,command=lambda: nhan_phim('35000'))
    bt1.grid(row=1, column=0)
    bt2 = Button(window, text=drink[1], width=15, height=3,command=lambda: nhan_phim('39000'))
    bt2.grid(row=1, column=1)
    bt3 = Button(window, text=drink[2], width=15, height=3,command=lambda: nhan_phim('39000'))
    bt3.grid(row=1, column=2)
    bt4 = Button(window, text=drink[3], width=15, height=3,command=lambda: nhan_phim('32000'))
    bt4.grid(row=2, column=0)
    bt5 = Button(window, text=drink[4], width=15, height=3,command=lambda: nhan_phim('29000'))
    bt5.grid(row=2, column=1)
    bt6 = Button(window, text=drink[5], width=15, height=3,command=lambda: nhan_phim('25000'))
    bt6.grid(row=2, column=2)
    bt7 = Button(window, text=drink[6], width=15, height=3,command=lambda: nhan_phim('29000'))
    bt7.grid(row=3, column=0)
    bt8 = Button(window, text=drink[7], width=15, height=3,command=lambda: nhan_phim('25000'))
    bt8.grid(row=3, column=1)
    #tạo nút tính
    cong = Button(window, text='+', width=15, height=3, command=lambda: nhan_phim('+'))
    cong.grid(row=6, column=1)
    tru = Button(window, text='-', width=15, height=3, command=lambda: nhan_phim('-'))
    tru.grid(row=6, column=2)
    bang = Button(window, text='Tổng cộng', width=15, height=3,command=lambda: nhan_bang())
    bang.grid(row=7, column=3)
    btn_xoa = Button(window, text='AC', width=15, height=3, command=lambda: xoa())
    btn_xoa.grid(row=6, column=3)
    #tạo bảng giá
    lb1 = Label(window, text='Bảng giá các món hàng', width=45, height=3, font=('Arial Bold', 13))
    lb1.grid(row=0, column=6, columnspan=4)
    lb2 = Label(window, text='Trà sữa:', width=15, height=3, font=('Arial Bold', 13))
    lb2.grid(row=1, column=6)
    lb3 = Label(window, text='Tra thanh đào:', width=15, height=3, font=('Arial Bold', 13))
    lb3.grid(row=2, column=6)
    lb4 = Label(window, text='Trà thạch đào:', width=15, height=3, font=('Arial Bold', 13))
    lb4.grid(row=3, column=6)
    lb5 = Label(window, text='Cà phê phin:', width=15, height=3, font=('Arial Bold', 13))
    lb5.grid(row=4, column=6)
    lb6 = Label(window, text='Ca phê sữa:', width=15, height=3, font=('Arial Bold', 13))
    lb6.grid(row=1, column=9)
    lb7 = Label(window, text='Ca phê đen:', width=15, height=3, font=('Arial Bold', 13))
    lb7.grid(row=2, column=9)
    lb8 = Label(window, text='Bạc sỉu:', width=15, height=3, font=('Arial Bold', 13))
    lb8.grid(row=3, column=9)
    lb9 = Label(window, text='Trà vải:', width=15, height=3, font=('Arial Bold', 13))
    lb9.grid(row=4, column=9)
    #BẢNG GIÁ TIỀN TỪNG MÓN
    lb10 = Label(window, text='35000vnd', width=15, height=3, font=('Arial', 13))
    lb10.grid(row=1, column=7)
    lb11 = Label(window, text='39000vnd', width=15, height=3, font=('Arial', 13))
    lb11.grid(row=2, column=7)
    lb12 = Label(window, text='39000vnd', width=15, height=3, font=('Arial', 13))
    lb12.grid(row=3, column=7)
    lb13 = Label(window, text='32000vnd', width=15, height=3, font=('Arial', 13))
    lb13.grid(row=4, column=7)
    lb14 = Label(window, text='29000vnd', width=15, height=3, font=('Arial', 13))
    lb14.grid(row=1, column=10)
    lb15 = Label(window, text='25000vnd', width=15, height=3, font=('Arial', 13))
    lb15.grid(row=2, column=10)
    lb16 = Label(window, text='29000vnd', width=15, height=3, font=('Arial', 13))
    lb16.grid(row=3, column=10)
    lb17 = Label(window, text='29000vnd', width=15, height=3, font=('Arial', 13))
    lb17.grid(row=4, column=10)
    #tạo option menu    
    opt1 = OptionMenu(window, quan[0], *optionList)
    opt1.grid(row=1, column=8)
    opt2 = OptionMenu(window, quan[1], *optionList)
    opt2.grid(row=2, column=8)
    opt3 = OptionMenu(window, quan[2], *optionList)
    opt3.grid(row=3, column=8)
    opt4 = OptionMenu(window, quan[3], *optionList)
    opt4.grid(row=4, column=8)
    opt5 = OptionMenu(window, quan[4], *optionList)
    opt5.grid(row=1, column=11)
    opt6 = OptionMenu(window, quan[5], *optionList)
    opt6.grid(row=2, column=11)
    opt7 = OptionMenu(window, quan[6], *optionList)
    opt7.grid(row=3, column=11)
    opt8 = OptionMenu(window, quan[7], *optionList)
    opt8.grid(row=4, column=11)
    btn_add = Button(window, text = 'Add a new invoice', command=add_invoice)
    btn_add.grid(row=5, column=9)
    btn_view_table = Button(window, text = 'View all invoices', command=obj.create)
    btn_view_table.grid(row=5, column=10)
    window.mainloop()