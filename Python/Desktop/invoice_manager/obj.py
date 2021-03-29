from tkinter import *
import openpyxl
class Table:
    def __init__ (self,window):
        self.data= []
        self.element= []
        self.window = window 
    def set_data(self,data):
        self.__data= data
    def get_data(self):
        return self.__data
    def print(self):
        for i in range(len(self.__data)):
            for j in range (len(self.__data[0])):
                line = []
                label = Label(self.window, width=20, text=str(self.__data[i][j]), fg='black',font=('Counrier New',12))
                label.grid(row=i + 1,column=j)
                line.append(label)
            self.element.append(line)
    def search(self,name='',place=""):
        result=[]
        for i in range(1, len(self.__data) - 1):
            if str(self.__data[i][1]).find(name) != -1 or str(self.__data[i][3]).find(place) != -1:
                result.append(self.__data[i])
        return result
def get_data_from_excel(filename):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook['Sheet1']
    other = workbook['Sheet2']
    row = int(other.cell(1, 1).value)
    result = []
    result.append(['CustomerID', 'Item Name', 'Quantity', 'Branch'])
    for i in range(row - 1):
        line = []
        for j in range(4):
            line.append(worksheet.cell(i + 1, j + 1).value)
        result.append(line)
    workbook.close()
    return result
def search_from_key(key):
    data = table.search(key.get(), key.get())
    table.set_data(data)
    table.print()
def create():
    global window
    window = Tk()
    data = get_data_from_excel('Expenses01.xlsx')
    global table
    table = Table(window)
    table.set_data(data)
    table.print()
    window.mainloop()